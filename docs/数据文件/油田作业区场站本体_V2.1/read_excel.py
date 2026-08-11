#!/usr/bin/env python3
"""读取三个 Excel 并导出为 JSON，避免 GBK 乱码"""

import openpyxl, json, os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

def read_sheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cleaned = []
            for v in row:
                if v is None:
                    cleaned.append("")
                elif isinstance(v, float):
                    cleaned.append(v)
                else:
                    cleaned.append(str(v))
            rows.append(dict(zip(headers, cleaned)))
        result[name] = {"headers": headers, "rows": rows, "count": len(rows)}
    return result

# 实体清单
entities = read_sheet('实体清单.xlsx')
with open('excel_entities.json', 'w', encoding='utf-8') as f:
    json.dump(entities, f, ensure_ascii=False, indent=2)

# 关系矩阵
relations = read_sheet('关系矩阵.xlsx')
with open('excel_relations.json', 'w', encoding='utf-8') as f:
    json.dump(relations, f, ensure_ascii=False, indent=2)

# 约束规则库
constraints = read_sheet('约束规则库.xlsx')
with open('excel_constraints.json', 'w', encoding='utf-8') as f:
    json.dump(constraints, f, ensure_ascii=False, indent=2)

print("OK: excel_entities.json, excel_relations.json, excel_constraints.json")

# 打印内容摘要
for fname, data in [("实体清单", entities), ("关系矩阵", relations), ("约束规则库", constraints)]:
    for sn, sd in data.items():
        print(f"\n=== {fname} — {sn} ({sd['count']}行) ===")
        print(f"  列: {sd['headers']}")
        for i, r in enumerate(sd['rows'][:8]):
            print(f"  [{i+1}] {json.dumps(r, ensure_ascii=False)}")
        if sd['count'] > 8:
            print(f"  ... 共 {sd['count']} 行")
