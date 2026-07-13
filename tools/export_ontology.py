#!/usr/bin/env python3
"""IO 服务器 DLAS 本体 → 交付物生成 + parse.db 入库"""
import sqlite3, json, time
from pathlib import Path

ROOT = Path(__file__).parent.parent
DOCS = ROOT / "docs"

# ═══════════════════════════════════════
# 1. 生成 Excel 交付物
# ═══════════════════════════════════════
def make_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[!] openpyxl not installed, skipping xlsx")
        return

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A4A6E", end_color="1A4A6E", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # ── 实体清单 ──
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "实体清单"
    ws.append(["层", "类别", "实体名称", "数量/标识", "属性/说明", "数据源"])
    for row in [
        ["Data", "服务器", "IO 服务器", "11.66.12.131", "Win2016, E:\\IO ServerOnLine", "WinRM"],
        ["Data", "服务器", "pSpace 服务器", "11.66.12.130", "psAPI :8889, 力控6.0.1.9", "psAPISDK"],
        ["Data", "服务器", "Oracle 11g", "11.66.12.129:1521", "DQYTPROD, OLEDB", "ADO"],
        ["Data", "设备", "RTU (Modbus)", "191台", "11.248-250.x, 主动连接:53001", "CommBridge"],
        ["Data", "设备", "OPC DA 端点", "5台", "Kepware 4.x, CLSID:6E6170F0-...", "DCOM"],
        ["Data", "设备", "保护装置", "12种", "DSL-31A/DST-31A/DSB-31A/...", "Device.ini"],
        ["Data", "软件", "CommBridge.exe", "155KB", "MFC主框架, TCP网关(191 RTU)", "PDB分析"],
        ["Data", "软件", "psAPISDK.dll", "3.5MB", "pSpace SDK (3525导出)", "SDK头文件"],
        ["Data", "软件", "IOMan.exe", "299KB", "IO管理器 (8实例)", "wmic"],
        ["Data", "软件", "IoMonitor.exe", "487KB", "GUI监视+Oracle写库", "PE分析"],
        ["Data", "软件", "IoProject.exe", "228KB", "pSpace核心调度(不依赖psAPI!)", "PE分析"],
        ["Data", "软件", "IoCommit.exe", "249KB", "批量写Oracle ×7", "PE分析"],
        ["Data", "软件", "GPRSDLL.dll", "1.38MB", "GPRS/CDMA协议栈", "Ghidra"],
        ["Data", "配置", "Device.ini", "12种设备", "通道映射+ChangeData系数(8192标定)", "文件读取"],
        ["Data", "配置", "IoChannelCfg.ini", "通道配置", "CommBridge DEV_COUNT=3", "文件读取"],
        ["Data", "配置", "SqlFilSet.ini", "Oracle连接", "EXECUTECYC=1000ms, ADOCOUNT=1", "文件读取"],
        ["Data", "配置", "IoMonitor.ini", "提交时序", "CommitRealSpan=300ms", "文件读取"],
        ["Data", "配置", "OPCClientCfg.ini", "OPC配置", "Kepware ProgID→CLSID", "文件读取"],
        ["Data", "配置", "DeviceStruct.txt", "OPC参数", "17字段(IP/ProgID/CLSID/刷新)", "文件读取"],
        ["Logic", "协议", "CommBridge帧", "TCP:53001", "Seq+Flags+Len+Slave+Func+Data", "pcapng(95K)"],
        ["Logic", "协议", "DTU注册包", "0xAA+Slave+ASCII_ID+0x0D", "2211包, 413种ID", "pcapng"],
        ["Logic", "协议", "pSpace TCP", ":8889", "psAPI + ACE框架(3525导出)", "SDK+实测"],
        ["Logic", "协议", "Modbus TCP", ":502", "Func 0x03(读保持寄存器)", "pcapng"],
        ["Logic", "校验", "L1 帧匹配", "abs(expected-actual)≤2", "", "commbridge_server.py"],
        ["Logic", "校验", "L2 值范围", "电流0-500A/电压100-400V", "", "RANGES"],
        ["Logic", "校验", "L3 三相平衡", "(Imax-Imin)/Iavg<25%", "", "cross_validate"],
        ["Logic", "校验", "L4 历时一致", "delta<50%", "", "cross_validate"],
        ["Logic", "校验", "L5 Oracle对标", "|A-B|/A<1%", "", "Oracle对比"],
        ["Action", "数据流", "RTU→CommBridge→IoMon→Oracle", "TCP+DCOM", "实时采集链路", "实测"],
        ["Action", "数据流", "DCS→OPC→IOMan→pSpace→IoCommit", "DCOM:135", "OPC DA链路", "实测"],
        ["Action", "IPC", "IoProject→IOMan", "CreateProcess", "命令行参数传递", "wmic"],
        ["Action", "IPC", "IOMan→IoMonitor", "共享内存+WM_COPYDATA", "Global\\命名空间", "字符串分析"],
        ["Action", "采集", "路径① Python Mock", "本机:502:13500:18889", "离线零依赖", "已验证"],
        ["Action", "采集", "路径③ psAPISDK直连", "130:8889", "ReadList实时读取", "已验证✅"],
        ["Security", "认证", "WinRM", ":5985 NTLM", "administrator", "巡检脚本"],
        ["Security", "认证", "Oracle", ":1521 OLEDB", "DQYTPROD", "SqlFilSet"],
        ["Security", "认证", "pSpace", ":8889 psAPI", "admin/DQYTA11_pass", "实测"],
        ["Security", "认证", "DCOM", ":135 CoInitSecurity", "DCS侧授权", "OPC配置"],
    ]:
        ws.append(row)
    for cell in ws[1]: cell.font = header_font; cell.fill = header_fill
    ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40; ws.column_dimensions['F'].width = 15
    wb.save(DOCS / "实体清单.xlsx"); print(f"[OK] {DOCS / '实体清单.xlsx'} ({ws.max_row-1} entities)")

    # ── 关系矩阵 ──
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active; ws2.title = "关系矩阵"
    ws2.append(["源实体", "关系类型", "目标实体", "协议/机制", "方向", "说明"])
    relations = [
        ("RTU设备", "monitors(监测)", "CommBridge", "TCP:53001 DTU透传", "→", "191台RTU主动连接"),
        ("CommBridge", "forwards(转发)", "IoMonitor", "WM_COPYDATA IPC", "→", "实时数据推送到监视器"),
        ("IoMonitor", "persists(持久化)", "Oracle 11g", "OLEDB ADO", "→", "批量写入DQYTPROD"),
        ("IoProject", "spawns(启动)", "IOMan", "CreateProcess+命令行", "→", "传递设备ID列表"),
        ("IOMan", "collects(采集)", "OPC DA Server", "DCOM:135", "↔", "OPC DA标准协议"),
        ("IOMan", "subscribes(订阅)", "pSpace Server", "TCP:8889 psAPI", "↔", "标签订阅+实时读取"),
        ("IoMonitor", "manages(管理)", "IoCommit", "共享内存", "→", "批量提交控制(7实例)"),
        ("pSpace Server", "bridges(桥接)", "OPC DA Server", "OPC DA DCOM", "→", "数据源→实时数据库"),
        ("传感器", "senses(感知)", "保护装置", "Modbus RTU", "→", "Ia/Ib/Ic/Ua/Ub/Uc/P/cosφ/F"),
        ("保护装置", "installedAt(安装于)", "变电站", "物理安装", "→", "一次设备关联"),
        ("ChangeData系数", "converts(转换)", "原始值", "Y×C[i]", "→", "8192标定→工程值"),
        ("dgiot_lite", "replaces(替代)", "IoMonitor+IoCommit", "Oracle管线+MQTT", "→", "数据同步+推送"),
        ("dev_env.py", "simulates(模拟)", "全套131架构", "Python socket", "→", "7服务+56设备离线"),
        ("pspace_collector", "reads(读取)", "pSpace Server", "psAPISDK ctypes", "→", "自主Tag ID实时读取"),
    ]
    for row in relations: ws2.append(row)
    for cell in ws2[1]: cell.font = header_font; cell.fill = header_fill
    for c in ['A','B','C','D','E','F']: ws2.column_dimensions[c].width = 22
    ws2.column_dimensions['F'].width = 40
    wb2.save(DOCS / "关系矩阵.xlsx"); print(f"[OK] {DOCS / '关系矩阵.xlsx'} ({len(relations)} relations)")

    # ── 约束规则库 ──
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active; ws3.title = "约束规则库"
    ws3.append(["层级", "规则名称", "阈值/条件", "严重度", "执行动作", "出处"])
    rules = [
        ("Logic", "RTU轮询间隔", "1000ms", "warn", "超时标记离线", "IoMonitor"),
        ("Logic", "OPC刷新时间", "设备参数[164]", "info", "按配置刷新", "DeviceStruct"),
        ("Logic", "数据提交(实时)", "300ms", "info", "批量提交Oracle", "IoMonitor.ini"),
        ("Logic", "提交批次大小", "15000点", "info", "分批次提交", "IoMonitor.ini"),
        ("Logic", "SQL执行周期", "1000ms", "info", "定时执行SQL", "SqlFilSet.ini"),
        ("Logic", "CommBridge超时", "30s", "warn", "断开重连", "IoChannelCfg.ini"),
        ("Logic", "帧长度匹配(L1)", "abs(expected-actual)≤2", "danger", "丢弃坏帧", "commbridge_server.py"),
        ("Logic", "电流范围(L2)", "0-500A", "danger", "标记异常+告警", "RANGES"),
        ("Logic", "电压范围(L2)", "100-400V", "danger", "标记异常+告警", "RANGES"),
        ("Logic", "三相平衡(L3)", "(Imax-Imin)/Iavg<25%", "danger", "触发三相不平衡告警", "cross_validate"),
        ("Logic", "历时一致性(L4)", "delta<50%", "warn", "标记可疑值", "cross_validate"),
        ("Logic", "Oracle对标(L5)", "|A-B|/A<1%", "danger", "数据源异常告警", "Oracle对比"),
        ("Logic", "C0=170/8192", "Ia,Ib,Ic,Ua,Ub,Uc", "info", "原始值→工程值", "Device.ini"),
        ("Logic", "C5=2/8192", "F=50+Y×2/8192 Hz", "info", "频率转换", "Device.ini"),
        ("Action", "自动重启", "3次异常→重启", "critical", "自动重启commbridge_server", "loop_131_monitor.py"),
        ("Action", "平台自恢复", "健康检查失败→重启", "critical", "杀端口→重启uvicorn", "platform_auto_restart.py"),
        ("Security", "生产网只读", "禁止安装/重启/停服", "critical", "仅抓包+WinRM查询", "production-host-rules"),
        ("Security", "凭据保护", "密码不入库/不写死", "critical", "环境变量或配置加密", "安全审计"),
    ]
    for row in rules: ws3.append(row)
    for cell in ws3[1]: cell.font = header_font; cell.fill = header_fill
    ws3.column_dimensions['A'].width = 10; ws3.column_dimensions['B'].width = 28
    ws3.column_dimensions['C'].width = 30; ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 30; ws3.column_dimensions['F'].width = 20
    wb3.save(DOCS / "约束规则库.xlsx"); print(f"[OK] {DOCS / '约束规则库.xlsx'} ({len(rules)} rules)")

# ═══════════════════════════════════════
# 2. parse.db 入库
# ═══════════════════════════════════════
def insert_parse_db():
    db_path = ROOT / "data" / "parse.db"
    db = sqlite3.connect(str(db_path))
    db.execute("PRAGMA journal_mode=WAL")

    # ── Site ──
    db.execute("INSERT OR IGNORE INTO ontology_site VALUES (?,?,?,?)",
               ("site_dqyt", "大庆油田作业区", "11.66.12.0/24", json.dumps({
                   "io_server": "11.66.12.131", "pspace": "11.66.12.130:8889",
                   "oracle": "11.66.12.129:1521", "rtu_count": 191
               }, ensure_ascii=False)))

    # ── Gateway (CommBridge) ──
    db.execute("INSERT OR IGNORE INTO ontology_gateway VALUES (?,?,?,?)",
               ("gw_cb_131", "CommBridge DTU网关", "site_dqyt", json.dumps({
                   "port": 53001, "protocol": "DTU transparent Modbus",
                   "frame": "Seq+Flags+Len+Slave+Func+Data", "rtu_count": 191
               }, ensure_ascii=False)))

    # ── Channels ──
    channels = [
        ("ch_modbus_tcp", "Modbus TCP", "gw_cb_131", "connect", ":502"),
        ("ch_opc_da", "OPC DA DCOM", "gw_cb_131", "dcom", "172.23.9.3:135"),
        ("ch_pspace", "pSpace psAPI", "gw_cb_131", "connect", "11.66.12.130:8889"),
        ("ch_dtu", "DTU Transparent", "gw_cb_131", "listen", ":53001"),
        ("ch_oracle", "Oracle OLEDB", "gw_cb_131", "poll", "11.66.12.129:1521"),
    ]
    for ch_id, ch_name, gw, mode, addr in channels:
        db.execute("INSERT OR IGNORE INTO ontology_channel VALUES (?,?,?,?,?)",
                   (ch_id, ch_name, gw, mode, addr))

    # ── Devices (20 real + 12 types) ──
    real_devices = {
        "02012170058": "DSL-31A 线路断路器", "02105100097": "DST-31A 变压器差动",
        "02106290043": "DSB-31A 变压器后备", "02107010048": "电动机保护",
        "02110080020": "DGP-13 接地保护", "02204060100": "DSL-31A 断路器",
        "02204060111": "DST-31A 变压器差动", "02105110008": "DST-31A 变压器差动",
        "02106290052": "DSB-31A 变压器后备", "02106290085": "DSB-31A 变压器后备",
        "02107030091": "电动机保护", "02107190091": "电动机保护",
        "02110080028": "DSL-31A 断路器", "02110110045": "DST-31A 变压器差动",
        "02110120089": "DBPA-31A 备用电源", "02110150030": "DSB-31A 变压器后备",
        "02110150041": "DSB-31A 变压器后备", "02110150046": "DSB-31A 变压器后备",
        "02110160086": "电动机保护", "02111270058": "DBPA-31A 备用电源",
    }
    for did, dname in real_devices.items():
        db.execute("INSERT OR IGNORE INTO ontology_device VALUES (?,?,?,?,?)",
                   (did, dname, "ch_modbus_tcp", "MODBUS_RTU", json.dumps(
                       {"status": "online", "source": "IOMan_wmic"}, ensure_ascii=False)))

    # ── Points (typical channels per device type) ──
    points = [("Ia", "相电流", "A", "FLOAT"), ("Ib", "相电流B", "A", "FLOAT"),
              ("Ic", "相电流C", "A", "FLOAT"), ("Ua", "相电压", "V", "FLOAT"),
              ("Ub", "相电压B", "V", "FLOAT"), ("Uc", "相电压C", "V", "FLOAT"),
              ("P", "有功功率", "kW", "FLOAT"), ("Q", "无功功率", "kVar", "FLOAT"),
              ("cos", "功率因数", "", "FLOAT"), ("F", "频率", "Hz", "FLOAT")]
    for dev_id in list(real_devices.keys())[:5]:
        for pt_id, pt_name, unit, dtype in points:
            db.execute("INSERT OR IGNORE INTO ontology_point VALUES (?,?,?,?,?,?)",
                       (f"{dev_id}.{pt_id}", dev_id, pt_name, unit, dtype,
                        json.dumps({"coeff": "170/8192" if pt_id.startswith("I") or pt_id.startswith("U") else "1"})))

    # ── Constraints ──
    constraints = [
        ("C_L1_frame", "L1 帧长度匹配", "abs(expected-actual)≤2", "danger",
         "丢弃坏帧", "commbridge_server.py"),
        ("C_L2_range", "L2 值范围校验", "current:0-500A, voltage:100-400V, power:0-300kW",
         "danger", "标记异常+告警", "RANGES"),
        ("C_L3_balance", "L3 三相平衡", "(Imax-Imin)/Iavg<25%",
         "danger", "触发三相不平衡告警", "cross_validate"),
        ("C_L4_temporal", "L4 历时一致性", "delta<50%",
         "warn", "标记可疑值", "cross_validate"),
        ("C_L5_oracle", "L5 Oracle对标", "|A-B|/A<1%",
         "danger", "触发数据源异常告警", "Oracle对比"),
        ("C_coeff", "ChangeData转换", "Y × Coefficient[i], 标定8192",
         "info", "原始值→工程值", "Device.ini"),
        ("C_prod_readonly", "生产网只读", "禁止安装/重启/停服",
         "critical", "仅抓包+WinRM查询", "production-host-rules"),
        ("C_heartbeat", "CommBridge心跳超时", "30s",
         "warn", "断开+重连", "IoChannelCfg.ini"),
        ("C_commit_time", "IoCommit提交间隔", "300ms实时/500ms历史",
         "info", "批量提交Oracle", "IoMonitor.ini"),
    ]
    for cid, cname, rule, severity, action, source in constraints:
        db.execute("INSERT OR IGNORE INTO ontology_constraint VALUES (?,?,?,?,?,?,?)",
                   (cid, cname, "Logic", rule, severity, action, source))

    # ── DataSources ──
    db.execute("INSERT OR IGNORE INTO ontology_datasource VALUES (?,?,?,?,?,?)",
               ("ds_pspace", "pSpace 6.0", "11.66.12.130:8889", "psAPI",
                "admin", json.dumps({"sdk_version": "6.0.1.9", "exports": 3525,
                 "connect_fn": "psAPI_Server_Connect", "read_fn": "psAPI_Real_ReadList",
                 "handle_type": "uint16"}, ensure_ascii=False)))
    db.execute("INSERT OR IGNORE INTO ontology_datasource VALUES (?,?,?,?,?,?)",
               ("ds_oracle", "Oracle 11g", "11.66.12.129:1521", "OLEDB",
                "DQYTPROD", json.dumps({"tables": ["TAGPAR","SYS_POINTRELATION",
                 "PROJECT_IODATASOURCE"]}, ensure_ascii=False)))
    db.execute("INSERT OR IGNORE INTO ontology_datasource VALUES (?,?,?,?,?,?)",
               ("ds_opc", "Kepware OPC DA", "172.23.9.3:135", "DCOM",
                "", json.dumps({"clsid": "6E6170F0-FF2D-11D2-8087-00105AA8F840",
                 "version": "4.x"}, ensure_ascii=False)))

    # ── Relations ──
    rels = [
        ("rel_rtu_cb", "02012170058", "gw_cb_131", "monitors", "RTU→CommBridge"),
        ("rel_cb_imon", "gw_cb_131", "site_dqyt", "forwards", "CommBridge→IoMonitor"),
        ("rel_ip_ioman", "gw_cb_131", "02012170058", "spawns", "IoProject→IOMan"),
        ("rel_ioman_pspace", "02012170058", "ds_pspace", "subscribes", "IOMan→pSpace"),
        ("rel_pspace_opc", "ds_pspace", "ds_opc", "bridges", "pSpace→OPC DA"),
        ("rel_imon_oracle", "site_dqyt", "ds_oracle", "persists", "IoMonitor→Oracle"),
        ("rel_dgiot_replace", "dgiot_lite", "site_dqyt", "replaces", "dgiot替代IoMonitor"),
    ]
    for rid, src, tgt, label, desc in rels:
        db.execute("INSERT OR IGNORE INTO Relation VALUES (?,?,?,?,?)",
                   (rid, src, tgt, label, desc))

    db.commit()

    # Verify
    counts = {}
    for table in ["ontology_site","ontology_gateway","ontology_channel","ontology_device",
                  "ontology_point","ontology_constraint","ontology_datasource","Relation"]:
        cnt = db.execute(f"SELECT count(*) FROM {table}").fetchone()[0]
        counts[table] = cnt
    db.close()

    print(f"[OK] parse.db updated:")
    for t, c in counts.items():
        print(f"  {t}: {c} rows")

# ═══════════════════════════════════════
# Main
# ═══════════════════════════════════════
if __name__ == "__main__":
    print("=== 本体交付物生成 ===")
    make_xlsx()
    print()
    insert_parse_db()
    print()
    print("Done. 交付物:")
    print(f"  {DOCS / '实体清单.xlsx'}")
    print(f"  {DOCS / '关系矩阵.xlsx'}")
    print(f"  {DOCS / '约束规则库.xlsx'}")
    print(f"  {ROOT / 'data' / 'parse.db'} (实时本体)")
